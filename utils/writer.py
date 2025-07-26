"""Excel writer for nurse scheduling results."""
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet


from .constants import TEMPLATE_PATH


def fill_shift_cells(ws: Worksheet, assignments: pd.DataFrame) -> None:
    """Fill each cell in the template sheet with the shift assignments.

    This implementation mimics the manual Excel workflow by first reading
    the nurse names already present in the template and then writing shift
    codes only for matching rows.  The template is assumed to have nurse
    names starting from row 6 in column A and day columns starting from
    column C.

    Parameters
    ----------
    ws : Worksheet
        Target worksheet ("シフト表").
    assignments : pandas.DataFrame
        DataFrame indexed by nurse names with columns day_1, day_2, ...
    """

    # Template rows containing nurse names, e.g. A6:A19
    start_row = 6
    nurse_rows = list(range(start_row, start_row + len(assignments.index)))
    nurse_names_in_excel = [ws.cell(row=r, column=1).value for r in nurse_rows]

    # Day columns in the result (max 31 days). Excel columns start at C (index 3)
    col_offset = 3
    date_cols = assignments.columns.tolist()[:31]

    for nurse in assignments.index:
        if nurse not in nurse_names_in_excel:
            # Skip nurses not present in the template
            continue
        row_idx = nurse_names_in_excel.index(nurse) + start_row
        for j, day_col in enumerate(date_cols):
            shift = assignments.at[nurse, day_col]
            ws.cell(row=row_idx, column=col_offset + j, value=shift)


def write_to_excel(
    df_result: pd.DataFrame,
    output_path: Path | str,
    template_path: Path | str = TEMPLATE_PATH,
) -> None:
    """Write the optimized schedule to an Excel file based on a template.

    Parameters
    ----------
    df_result : pandas.DataFrame
        DataFrame of shift assignments where index is nurse names and
        columns correspond to days (day_1, day_2, ...).
    output_path : str
        Path to save the resulting Excel file.
    template_path : str, optional
        Path to the Excel template, by default ``data/shift_template.xlsx``.
    """
    wb = load_workbook(str(template_path))
    ws = wb["シフト表"]
    fill_shift_cells(ws, df_result)
    wb.save(str(output_path))
